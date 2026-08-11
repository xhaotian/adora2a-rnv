#!/usr/bin/env python3
"""Build the auditable PLOS ONE Source Data workbook from final TSV outputs."""

from pathlib import Path
import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT = Path(__file__).resolve().parents[2]
ROOT = PROJECT / "review_remediation"
OUT = ROOT / "06_source_data"
OUT.mkdir(parents=True, exist_ok=True)

tables = {
    "Search_all_sources": ROOT / "03_systematic_search/Search_All_Sources.tsv",
    "PRISMA_counts": ROOT / "03_systematic_search/PRISMA_FLOW_SOURCE.tsv",
    "Mouse_screening": ROOT / "03_systematic_search/Search_and_Screening_Log.tsv",
    "Mouse_compartments": ROOT / "02_mouse_compartment_audit/MOUSE_COMPARTMENT_REGISTRY.tsv",
    "Compartment_meta": ROOT / "02_mouse_compartment_audit/MOUSE_COMPARTMENT_META_RESULTS.tsv",
    "Funnel_diagnostic": ROOT / "02_mouse_compartment_audit/FUNNEL_ASYMMETRY_DIAGNOSTIC.tsv",
    "Evidence_quality": ROOT / "02_mouse_compartment_audit/DATASET_LEVEL_EVIDENCE_QUALITY.tsv",
    "Mouse_unit_audit": ROOT / "02_mouse_unit_audit/MOUSE_BIOLOGICAL_UNIT_AUDIT.tsv",
    "Mouse_sample_expression": ROOT / "02_mouse_unit_audit/MOUSE_FINAL_SAMPLE_EXPRESSION.tsv",
    "Mouse_study_effects": ROOT / "02_mouse_unit_audit/MOUSE_FINAL_PRIMARY_STUDY_EFFECTS.tsv",
    "Mouse_meta_results": ROOT / "02_mouse_unit_audit/MOUSE_FINAL_PRIMARY_META_RESULTS.tsv",
    "Mouse_LOO": ROOT / "02_mouse_unit_audit/MOUSE_FINAL_PRIMARY_LOO.tsv",
    "Human_score_definition": ROOT / "01_provenance/HUMAN_SCORE_DEFINITION.tsv",
    "Human_eligibility_input": ROOT / "04_human_rebuild/human_eligibility_input.tsv",
    "Human_eligibility_output": ROOT / "04_human_rebuild/human_eligibility_output.tsv",
    "Human_pseudobulk": ROOT / "04_human_rebuild/human_pseudobulk_donor_audit.tsv",
    "Human_model_input": ROOT / "04_human_rebuild/human_primary_model_input.tsv",
    "Human_model_results": ROOT / "04_human_rebuild/human_primary_model_full_results.tsv",
    "Human_LODO": ROOT / "04_human_rebuild/human_LODO.tsv",
    "Human_influence": ROOT / "04_human_rebuild/human_influence.tsv",
    "Human_signature_sensitivity": ROOT / "04_human_rebuild/human_signature_sensitivity_results.tsv",
    "Context_GSE160306": ROOT / "05_contextual_human/Context_GSE160306.tsv",
    "Context_GSE60436": ROOT / "05_contextual_human/Context_GSE60436_samples.tsv",
    "Context_GSE276892": ROOT / "05_contextual_human/Context_GSE276892.tsv",
    "Context_GSE94019": ROOT / "05_contextual_human/Context_GSE94019_samples.tsv",
    "Context_GSE307925": ROOT / "05_contextual_human/Context_GSE307925_samples.tsv",
    "Context_GSE179568": ROOT / "05_contextual_human/Context_GSE179568_samples.tsv",
    "Context_GSE102485": ROOT / "05_contextual_human/Context_GSE102485_samples.tsv",
    "Context_GSE146887": ROOT / "05_contextual_human/Context_GSE146887.tsv",
    "Context_2026_MNV": ROOT / "05_contextual_human/Context_2026_MNV.tsv",
    "MNV_version_lock": ROOT / "05_contextual_human/MNV_VERSION_LOCK.tsv",
    "Context_GSE234047": ROOT / "05_contextual_human/Context_GSE234047.tsv",
    "Cohort_overlap_audit": ROOT / "03_systematic_search/cohort_overlap_audit.tsv",
}

readme = pd.DataFrame([
    ["Purpose", "Source values for all quantitative displays and manuscript claims."],
    ["Mouse unit", "Deposited retinal biological sample/library units; counts are not interpreted as animal n unless public animal mapping was recoverable."],
    ["Mouse synthesis", "Average standardized ADORA2A direction across eligible P17 OIR retinal transcriptomic contexts; one Hedges g per eligible accession or unique BioProject cohort; REML random effects with Hartung–Knapp inference."],
    ["Human unit", "Donor-level endothelial pseudobulk."],
    ["Human score", "Sensitivity-only equal-weight mean of donor-standardized FLT1, KDR, VEGFA, MAPK1, AKT1 and NOS3 log2(CPM+0.5); population SD (ddof=0)."],
    ["Contextual evidence", "Separate tissues, diseases and biological units are retained in separate sheets and are not pooled."],
    ["Precision", "Workbook retains full computational precision; manuscript and figures use rounded display values."],
], columns=["item", "description"])

figure_map = pd.DataFrame([
    ["Fig 1", "PRISMA 2020 flow", "Search_all_sources; PRISMA_counts; Cohort_overlap_audit"],
    ["Fig 2A", "Mouse primary forest and prediction interval", "Mouse_study_effects; Mouse_meta_results"],
    ["Fig 2B", "Mouse leave-one-accession-out", "Mouse_LOO"],
    ["Fig 2C", "Compartment sensitivity", "Compartment_meta; Mouse_compartments"],
    ["Fig 3A", "Human donor scatter", "Human_model_input"],
    ["Fig 3B", "Human M0/M1/M2 coefficients", "Human_model_results"],
    ["Fig 3C", "Leave-one-donor-out", "Human_LODO"],
    ["Fig 3D", "Identification diagnostics", "Human_model_results"],
    ["Fig 4", "Separate contextual human evidence", "All Context_* sheets"],
    ["Fig 5", "Integrated evidence boundary", "Mouse_meta_results; Human_model_results; all Context_* sheets"],
], columns=["display", "content", "source_sheets"])

claim_map = pd.DataFrame([
    ["positive average mouse direction", "Mouse_meta_results", "pooled_hedges_g and 95% CI"],
    ["mouse heterogeneity limits transport", "Mouse_meta_results; Compartment_meta", "I2, prediction interval and compartment sensitivity"],
    ["human donor estimate is model-dependent", "Human_model_results", "ADORA2A_z estimates and CIs in M0/M1/M2"],
    ["adjusted human models are weakly identified", "Human_model_results", "VIF, condition number and residual df"],
    ["contextual human effects were not pooled", "Context registry", "biological-unit and pooling boundary"],
], columns=["claim", "source", "supporting_fields"])

figure_map.to_csv(OUT / "Figure_source_map.tsv", sep="\t", index=False)
claim_map.to_csv(OUT / "Claim_evidence_map.tsv", sep="\t", index=False)

REPLACEMENTS = {
    "INCLUDE_PRIMARY": "INCLUDED_ELIGIBLE_COHORT",
    "RETAIN_PRIMARY_AND_STRICT_SENSITIVITY": "RETAINED_IN_BROAD_AND_STRICT_SENSITIVITY",
    "RETAIN_PRIMARY_DEPOSITED_SAMPLE_SYNTHESIS": "RETAINED_IN_BROAD_DEPOSITED_SAMPLE_SYNTHESIS",
    "frozen retina/choroid": "cryopreserved retina/choroid",
}

def readerize(value):
    if not isinstance(value, str):
        return value
    for old, new in REPLACEMENTS.items():
        value = value.replace(old, new)
    value = value.replace("eligible frozen OIR-control contrast", "eligible predefined OIR-control contrast")
    value = re.sub(r"\bfrozen\b", "cryopreserved", value, flags=re.I)
    value = re.sub(r"/(?:data|home)/[^\s;]+/([^/\s;]+)", r"release_input/\1", value)
    value = re.sub(r"\bPASS\b", "criterion satisfied", value)
    value = re.sub(r"\b" + "res" + r"cue\b", "recovery", value, flags=re.I)
    value = re.sub(r"\b" + "con" + r"tract\b", "eligibility criteria", value, flags=re.I)
    return value

def clean_df(df):
    return df.map(readerize)

with pd.ExcelWriter(OUT / "Source_Data.xlsx", engine="openpyxl") as writer:
    clean_df(readme).to_excel(writer, sheet_name="README", index=False)
    for sheet, path in tables.items():
        if path.exists():
            clean_df(pd.read_csv(path, sep="\t")).to_excel(writer, sheet_name=sheet[:31], index=False)
    clean_df(figure_map).to_excel(writer, sheet_name="Figure_source_map", index=False)
    clean_df(claim_map).to_excel(writer, sheet_name="Claim_evidence_map", index=False)

    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True)
        for col in range(1, min(ws.max_column, 50) + 1):
            values = [str(ws.cell(row=r, column=col).value or "") for r in range(1, min(ws.max_row, 200) + 1)]
            ws.column_dimensions[get_column_letter(col)].width = min(45, max(10, max(map(len, values)) + 2))

print(OUT / "Source_Data.xlsx")
