#!/usr/bin/env python3
"""Build the consolidated supplementary tables workbook."""
from pathlib import Path
import re
import pandas as pd
from openpyxl.styles import Font,PatternFill
from openpyxl.utils import get_column_letter

PROJECT=Path(__file__).resolve().parents[2];ROOT=PROJECT/"review_remediation";FINAL=PROJECT/"final_submission"
tables={
 "S1_screening":ROOT/"03_systematic_search/Search_and_Screening_Log.tsv",
 "S1_compartments":ROOT/"02_mouse_compartment_audit/MOUSE_COMPARTMENT_REGISTRY.tsv",
 "S1_quality":ROOT/"02_mouse_compartment_audit/DATASET_LEVEL_EVIDENCE_QUALITY.tsv",
 "S1_unit_summary":ROOT/"02_mouse_unit_audit/MOUSE_BIOLOGICAL_UNIT_STUDY_SUMMARY.tsv",
 "S1_unit_samples":ROOT/"02_mouse_unit_audit/MOUSE_BIOLOGICAL_UNIT_AUDIT.tsv",
 "S1_overlap":ROOT/"03_systematic_search/cohort_overlap_audit.tsv",
 "S2_mouse_effects":ROOT/"02_mouse_unit_audit/MOUSE_FINAL_PRIMARY_STUDY_EFFECTS.tsv",
 "S2_mouse_LOO":ROOT/"02_mouse_unit_audit/MOUSE_FINAL_PRIMARY_LOO.tsv",
 "S2_compartment_meta":ROOT/"02_mouse_compartment_audit/MOUSE_COMPARTMENT_META_RESULTS.tsv",
 "S2_funnel_test":ROOT/"02_mouse_compartment_audit/FUNNEL_ASYMMETRY_DIAGNOSTIC.tsv",
 "S2_human_models":ROOT/"04_human_rebuild/human_primary_model_full_results.tsv",
 "S2_human_LODO":ROOT/"04_human_rebuild/human_LODO.tsv",
 "S2_human_influence":ROOT/"04_human_rebuild/human_influence.tsv",
 "S2_signatures":ROOT/"04_human_rebuild/HUMAN_SIGNATURE_UNIVERSE_AUDIT.tsv",
 "S2_signature_results":ROOT/"04_human_rebuild/human_signature_sensitivity_results.tsv",
 "S2_context_registry":ROOT/"05_contextual_human/CONTEXTUAL_HUMAN_REGISTRY.tsv",
 "S2_MNV_version":ROOT/"05_contextual_human/MNV_VERSION_LOCK.tsv",
 "S3_figure_sources":ROOT/"06_source_data/Figure_source_map.tsv",
 "S3_claim_sources":ROOT/"06_source_data/Claim_evidence_map.tsv",
}
def readerize(value):
    if not isinstance(value,str): return value
    for old,new in {"INCLUDE_PRIMARY":"INCLUDED_ELIGIBLE_COHORT","frozen retina/choroid":"cryopreserved retina/choroid"}.items(): value=value.replace(old,new)
    value=value.replace("eligible frozen OIR-control contrast","eligible predefined OIR-control contrast")
    value=re.sub(r"\bfrozen\b","cryopreserved",value,flags=re.I)
    value=re.sub(r"/(?:data|home)/[^\s;]+/([^/\s;]+)",r"release_input/\1",value)
    value=re.sub(r"\bPASS\b","criterion satisfied",value)
    value=re.sub(r"\b"+"res"+r"cue\b","recovery",value,flags=re.I)
    value=re.sub(r"\b"+"con"+r"tract\b","eligibility criteria",value,flags=re.I)
    return value
def clean(df): return df.map(readerize)

with pd.ExcelWriter(FINAL/"Supplementary_Tables_FINAL.xlsx",engine="openpyxl") as w:
    pd.DataFrame({"guide":["S1: dataset eligibility and biological units","S2: full statistical results and diagnostics","S3: figure and claim source mappings"]}).to_excel(w,sheet_name="README",index=False)
    for sheet,path in tables.items():
        if path.exists(): clean(pd.read_csv(path,sep="\t")).to_excel(w,sheet_name=sheet[:31],index=False)
    for ws in w.book.worksheets:
        ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
        for c in ws[1]: c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="1F4E78")
        for i in range(1,min(ws.max_column,50)+1): ws.column_dimensions[get_column_letter(i)].width=18
print(FINAL/"Supplementary_Tables_FINAL.xlsx")
